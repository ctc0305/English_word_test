from openpyxl.styles import PatternFill

def load_xlsx_file(fp, spliter, start, final, sheets):
    word = []
    numbers = []
    total_word_number = 0
    for i in range(start - 1, final):
        sheet = fp[sheets[i]]
        numbers.append(sheet.max_row)
        for j in range(numbers[i - start + 1]):
            try:
                word.append(sheet.cell(j+1, 1).value.split(spliter))
            except:
                print("工作表含無效的儲存格!")
                quit()
        total_word_number += numbers[i - start + 1]
    return word, numbers, total_word_number

def mark_cells(color, fp, index, number, sheets, start):
    order = 0
    sheet_name = ''
    for i in range(start - 1, len(sheets)):#尋找要標記的位置
        if order + number[i - start + 1] <= index:
            order += number[i - start + 1]
        else:
            sheet_name = sheets[i]
            break
    sheet = fp[sheet_name]
    if color:#如果答案正確填綠色，反之則填紅色
        sheet.cell(index - order + 1, 1).fill = PatternFill("solid", fgColor = "00FF00")
    else:
        sheet.cell(index - order + 1, 1).fill = PatternFill("solid", fgColor = "FF0000")
    
